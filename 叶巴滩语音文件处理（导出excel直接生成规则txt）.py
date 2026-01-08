import tkinter as tk
from tkinter import filedialog, messagebox
import pandas as pd
import openpyxl
import os


# ================= 规则加载 =================

def load_replace_rules(rule_file):
    rules = {}
    with open(rule_file, "r", encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                old, new = line.split("=", 1)
                rules[old] = new
    return rules


# ================= 核心整合处理逻辑 =================

def process_excel_all_in_one(
    input_excel,
    output_txt,
    rule_file,
    gui_prefix
):
    # ---------- 1. 读取原始 Excel ----------
    wb = openpyxl.load_workbook(input_excel)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]

    def col_index(name):
        if name not in headers:
            raise ValueError(f"Excel 缺少列：{name}")
        return headers.index(name)

    idx_id = col_index("序号")
    idx_desc = col_index("描述")
    idx_name = col_index("名称")

    # ---------- 2. 找到「名称列标红」的行 ----------
    valid_rows = []

    for row in ws.iter_rows(min_row=2):
        name_cell = row[idx_name]
        font_color = name_cell.font.color

        if font_color and font_color.type == "rgb":
            rgb = font_color.rgb
            if rgb and rgb[-6:].upper() == "FF0000":
                valid_rows.append(row)

    if not valid_rows:
        raise ValueError("未找到任何「名称列为红色」的行")

    # ---------- 3. 构造等价于“预处理 Excel”的数据 ----------
    data = {
        "序号": [],
        "前缀": [],
        "描述": [],
        "名称": []
    }

    for row in valid_rows:
        data["序号"].append(row[idx_id].value)
        data["前缀"].append("")          # 原程序1里的前缀列
        data["描述"].append(row[idx_desc].value)
        data["名称"].append(row[idx_name].value)

    df = pd.DataFrame(data)

    # ---------- 4. 拼接完整文本（GUI 前缀 + 前缀列 + 名称） ----------
    df["原始文本"] = (
        gui_prefix +
        df["前缀"].astype(str) +
        df["名称"].astype(str)
    )

    # ---------- 5. 加载并执行规则替换 ----------
    replace_rules = load_replace_rules(rule_file)

    df["最终文本"] = df["原始文本"].replace(
        replace_rules,
        regex=True
    )

    # ---------- 6. 描述加 .wav ----------
    df["描述_处理后"] = df["描述"].astype(str) + ".wav"

    # ---------- 7. 导出 txt ----------
    out_df = df[["序号", "最终文本", "描述_处理后"]]

    out_df.to_csv(
        output_txt,
        sep=" ",
        index=False,
        header=False,
        encoding="ANSI"
    )


# ================= GUI =================

def choose_input():
    path = filedialog.askopenfilename(
        title="选择输入 Excel",
        filetypes=[("Excel 文件", "*.xlsx")]
    )
    if path:
        entry_input.delete(0, tk.END)
        entry_input.insert(0, path)


def choose_output_dir():
    path = filedialog.askdirectory(title="选择输出文件夹")
    if path:
        entry_out_dir.delete(0, tk.END)
        entry_out_dir.insert(0, path)


def choose_rule():
    path = filedialog.askopenfilename(
        title="选择规则文件",
        filetypes=[("规则文件", "*.txt")]
    )
    if path:
        entry_rule.delete(0, tk.END)
        entry_rule.insert(0, path)


def run():
    input_excel = entry_input.get().strip()
    output_dir = entry_out_dir.get().strip()
    output_name = entry_out_name.get().strip()
    rule_file = entry_rule.get().strip()
    gui_prefix = entry_prefix.get().strip()

    if not input_excel:
        messagebox.showerror("错误", "请选择输入 Excel")
        return
    if not output_dir:
        messagebox.showerror("错误", "请选择输出文件夹")
        return
    if not output_name:
        messagebox.showerror("错误", "请输入输出文件名")
        return
    if not rule_file:
        messagebox.showerror("错误", "请选择规则文件")
        return

    if not output_name.lower().endswith(".txt"):
        output_name += ".txt"

    output_txt = os.path.join(output_dir, output_name)

    try:
        process_excel_all_in_one(
            input_excel=input_excel,
            output_txt=output_txt,
            rule_file=rule_file,
            gui_prefix=gui_prefix
        )
        messagebox.showinfo("完成", f"生成成功：\n{output_txt}")
    except Exception as e:
        messagebox.showerror("失败", str(e))


# ================= 主窗口 =================

root = tk.Tk()
root.title("Excel → TXT 点表一体化处理工具")
root.geometry("700x350")
root.resizable(False, False)

tk.Label(root, text="输入 Excel：").grid(row=0, column=0, padx=30, pady=12, sticky="e")
entry_input = tk.Entry(root, width=58)
entry_input.grid(row=0, column=1)
tk.Button(root, text="选择", command=choose_input).grid(row=0, column=2)

tk.Label(root, text="输出文件夹：").grid(row=1, column=0, padx=30, pady=12, sticky="e")
entry_out_dir = tk.Entry(root, width=58)
entry_out_dir.grid(row=1, column=1)
tk.Button(root, text="选择", command=choose_output_dir).grid(row=1, column=2)

tk.Label(root, text="输出文件名：").grid(row=2, column=0, padx=30, pady=12, sticky="e")
entry_out_name = tk.Entry(root, width=58)
entry_out_name.grid(row=2, column=1, sticky="w")
tk.Label(root, text="（可不写 .txt）").grid(row=2, column=2, sticky="w")

tk.Label(root, text="规则文件：").grid(row=3, column=0, padx=30, pady=12, sticky="e")
entry_rule = tk.Entry(root, width=40)
entry_rule.grid(row=3, column=1,sticky="w")
tk.Button(root, text="选择", command=choose_rule).grid(row=3, column=2)

tk.Label(root, text="前缀（参与替换）：").grid(row=4, column=0, padx=30, pady=12, sticky="e")
entry_prefix = tk.Entry(root, width=58)
entry_prefix.grid(row=4, column=1, sticky="w")

tk.Button(
    root,
    text="开始处理",
    width=24,
    height=2,
    command=run
).grid(row=5, column=1, pady=30)

root.mainloop()
