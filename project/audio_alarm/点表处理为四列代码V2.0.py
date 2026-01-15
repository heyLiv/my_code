import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import load_workbook, Workbook
import copy
import os


# ================== 核心处理逻辑 ==================

def process_excel(input_path, output_path, prefix_value):
    src_wb = load_workbook(input_path)
    src_ws = src_wb.active

    headers = [cell.value for cell in src_ws[1]]
    idx_id = headers.index("序号") + 1
    idx_desc = headers.index("描述") + 1
    idx_name = headers.index("名称") + 1

    dst_wb = Workbook()
    dst_ws = dst_wb.active
    dst_ws.title = "结果"

    new_headers = ["序号", "前缀", "描述", "名称"]
    for col, h in enumerate(new_headers, start=1):
        dst_ws.cell(row=1, column=col, value=h)

    for row in range(2, src_ws.max_row + 1):
        # 序号
        c = src_ws.cell(row=row, column=idx_id)
        nc = dst_ws.cell(row=row, column=1, value=c.value)
        nc.font = copy.copy(c.font)

        # 前缀
        dst_ws.cell(row=row, column=2, value=prefix_value)

        # 描述
        c = src_ws.cell(row=row, column=idx_desc)
        nc = dst_ws.cell(row=row, column=3, value=c.value)
        nc.font = copy.copy(c.font)

        # 名称（保留标红）
        c = src_ws.cell(row=row, column=idx_name)
        nc = dst_ws.cell(row=row, column=4, value=c.value)
        nc.font = copy.copy(c.font)

    dst_wb.save(output_path)


# ================== GUI ==================

def choose_input_file():
    path = filedialog.askopenfilename(
        title="选择输入 Excel 文件",
        filetypes=[("Excel 文件", "*.xlsx")]
    )
    if path:
        input_entry.delete(0, tk.END)
        input_entry.insert(0, path)


def choose_output_dir():
    path = filedialog.askdirectory(title="选择输出文件夹")
    if path:
        output_dir_entry.delete(0, tk.END)
        output_dir_entry.insert(0, path)


def run_process():
    input_path = input_entry.get().strip()
    output_dir = output_dir_entry.get().strip()
    prefix = prefix_entry.get().strip()
    output_name = output_name_entry.get().strip()

    if not input_path:
        messagebox.showerror("错误", "请选择输入 Excel 文件")
        return
    if not output_dir:
        messagebox.showerror("错误", "请选择输出文件夹")
        return
    if not prefix:
        messagebox.showerror("错误", "请输入前缀名称")
        return
    if not output_name:
        messagebox.showerror("错误", "请输入输出文件名")
        return

    # 处理文件名
    if not output_name.lower().endswith(".xlsx"):
        output_name += ".xlsx"

    output_path = os.path.join(output_dir, output_name)

    try:
        process_excel(input_path, output_path, prefix)
        messagebox.showinfo(
            "完成",
            f"处理完成！\n\n生成文件：\n{output_path}"
        )
    except Exception as e:
        messagebox.showerror("失败", str(e))


# ================== 主窗口 ==================

root = tk.Tk()
root.title("Excel 点表整理工具")
root.geometry("640x300")
root.resizable(False, False)

# 输入文件
tk.Label(root, text="输入 Excel：").grid(row=0, column=0, padx=10, pady=12, sticky="e")
input_entry = tk.Entry(root, width=50)
input_entry.grid(row=0, column=1)
tk.Button(root, text="选择文件", command=choose_input_file).grid(row=0, column=2, padx=5)

# 输出文件夹
tk.Label(root, text="输出文件夹：").grid(row=1, column=0, padx=10, pady=12, sticky="e")
output_dir_entry = tk.Entry(root, width=50)
output_dir_entry.grid(row=1, column=1)
tk.Button(root, text="选择文件夹", command=choose_output_dir).grid(row=1, column=2, padx=5)

# 前缀
tk.Label(root, text="前缀名称：").grid(row=2, column=0, padx=10, pady=12, sticky="e")
prefix_entry = tk.Entry(root, width=30)
prefix_entry.grid(row=2, column=1, sticky="w")

# 输出文件名
tk.Label(root, text="输出文件名：").grid(row=3, column=0, padx=10, pady=12, sticky="e")
output_name_entry = tk.Entry(root, width=30)
output_name_entry.grid(row=3, column=1, sticky="w")
tk.Label(root, text="（可不写 .xlsx）").grid(row=3, column=2, sticky="w")

# 按钮
tk.Button(
    root,
    text="开始处理",
    width=20,
    height=2,
    command=run_process
).grid(row=4, column=1, pady=20)

root.mainloop()
